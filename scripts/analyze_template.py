"""
양식 파일 전체 내용 추출기 — 상위 5행이 아닌 전체 시트 내용 추출
지침/예시/헤더 등 양식 내 모든 정보를 캡처하여 AI에 전달한다.
"""
import json, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_CELL_CHARS = 300
MAX_ROWS_PER_SHEET = 200
MAX_FORMULAS_PER_SHEET = 80
MAX_COMMENTS_PER_SHEET = 80
MAX_STYLE_HINTS_PER_SHEET = 120


def _cell_ref(row_idx, col_idx):
    return f"{get_column_letter(col_idx)}{row_idx}"


def _fill_color(cell):
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        return color.rgb
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def _font_hint(cell):
    hints = []
    if cell.font and cell.font.bold:
        hints.append("bold")
    if cell.alignment and cell.alignment.wrap_text:
        hints.append("wrap")
    return ",".join(hints) if hints else None


def analyze(template_path):
    wb_values = load_workbook(template_path, read_only=True, data_only=True)
    wb_meta = load_workbook(template_path, read_only=False, data_only=False)
    result = {
        "sheets": [],
        "total_sheets": len(wb_values.sheetnames),
    }

    for sheet_name in wb_values.sheetnames:
        ws = wb_values[sheet_name]
        ws_meta = wb_meta[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": [],
            "merged_cells": [str(rng) for rng in ws_meta.merged_cells.ranges],
            "formulas": [],
            "comments": [],
            "style_hints": [],
            "hidden_rows": [],
            "hidden_cols": [],
        }

        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_count >= MAX_ROWS_PER_SHEET:
                sheet_info["truncated"] = True
                break

            non_empty = {}
            for col_idx, val in enumerate(row, 1):
                if val is not None:
                    s = str(val).strip()
                    if s:
                        non_empty[col_idx] = s[:MAX_CELL_CHARS]

            if non_empty:
                sheet_info["rows"].append({"row": row_idx, "cells": non_empty})
                row_count += 1

        for row_idx, dim in ws_meta.row_dimensions.items():
            if dim.hidden:
                sheet_info["hidden_rows"].append(row_idx)
        for col_letter, dim in ws_meta.column_dimensions.items():
            if dim.hidden:
                sheet_info["hidden_cols"].append(col_letter)

        for row in ws_meta.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    if len(sheet_info["formulas"]) < MAX_FORMULAS_PER_SHEET:
                        sheet_info["formulas"].append({
                            "cell": cell.coordinate,
                            "formula": value[:MAX_CELL_CHARS],
                        })
                if cell.comment and len(sheet_info["comments"]) < MAX_COMMENTS_PER_SHEET:
                    sheet_info["comments"].append({
                        "cell": cell.coordinate,
                        "text": str(cell.comment.text).strip()[:MAX_CELL_CHARS],
                    })
                fill = _fill_color(cell)
                font = _font_hint(cell)
                if (fill or font) and cell.value not in (None, ""):
                    if len(sheet_info["style_hints"]) < MAX_STYLE_HINTS_PER_SHEET:
                        sheet_info["style_hints"].append({
                            "cell": cell.coordinate,
                            "value": str(cell.value).strip()[:120],
                            "fill": fill,
                            "font": font,
                        })

        result["sheets"].append(sheet_info)

    wb_values.close()
    wb_meta.close()
    return result


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--template", required=True)
    args = p.parse_args()
    print(json.dumps(analyze(args.template), ensure_ascii=False, indent=2))
