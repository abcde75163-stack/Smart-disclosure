"""
기술이전 실적 세부 현황 (29번 양식) 자동 생성

사용법:
  python3 extract_detail.py \
    --master "기술이전총정리.xlsx" \
    --year 2025 \
    --output "기술이전실적세부현황_2025.xlsx"
"""
import argparse, sys, os
sys.path.insert(0, os.path.dirname(__file__))
from common import *
from feedback import write_feedback_sheet
from openpyxl import Workbook

TECH_TYPE_TEXT = {
    1:'특허', '1':'특허', '특허':'특허',
    2:'실용신안', '2':'실용신안',
    3:'디자인', '3':'디자인',
    4:'상표', '4':'상표',
    5:'소프트웨어', '5':'소프트웨어',
    6:'노하우', '6':'노하우', '노하우':'노하우', '정보 및 노하우':'노하우',
    7:'기타', '7':'기타',
    8:'단위연구노하우', '8':'단위연구노하우',
    9:'기술자문', '9':'기술자문',
    10:'저작권', '10':'저작권', '저작권':'저작권',
}

TRADE_TYPE_TEXT = {
    0:'매매', '0':'매매', '양도(매매)':'매매', '매매':'매매', '일부양도':'매매',
    1:'전용실시', '1':'전용실시', '전용실시':'전용실시',
    2:'통상실시', '2':'통상실시', '통상실시':'통상실시',
    '독점적통상실시':'통상실시', '통상실시(독점)':'통상실시',
    3:'노하우', '3':'노하우',
    4:'기술제휴', '4':'기술제휴',
    5:'OEM/ODM',
    6:'M&A',
    7:'기타', '7':'기타', '기타':'기타',
    8:'단위연구노하우', '8':'단위연구노하우',
    9:'기술자문', '9':'기술자문', '기술자문':'기술자문',
    '저작권':'저작권',
}

def get_text(mapping, raw):
    if raw is None: return ''
    key = raw if raw in mapping else str(raw)
    return mapping.get(key, str(raw))

def run(master_path, year, output_path):
    print(f"📂 DB 로딩: {master_path}")
    all_rows = load_master_db(master_path)

    # 필터: 입금일이 year인 건
    filtered = filter_by_year(all_rows, year, mode='payment')
    print(f"  → {year}년 입금 건: {len(filtered)}건")

    import datetime
    filtered.sort(key=lambda r: r[1] if isinstance(r[1], (datetime.datetime, datetime.date))
                  else datetime.datetime(1900,1,1))

    wb = Workbook()
    ws = wb.active
    ws.title = "29. 기술이전 실적 세부 현황"

    headers = [
        '순번', '기술이전 계약관리번호', '기술이전 계약일', '기술이전 계약 당사자',
        '업체명', '사업자등록번호', '기관유형', '업종유형', '국내/국외', '국가',
        '지역구분', '계약명', '대표 발명자', '기술(또는 권리)유형',
        '지식재산권 등 이전기술 수', '지식재산권번호 및 노하우 명', '기술분야(6T)', '기술분류',
        '거래 유형', '계약기간', '제한사항', '제한사항 기타', '제한사항설명',
        '기술료 수취유형', '정액 기술료(원)', '경상기술료(조건)', '주식 수(주)', '액면가액(원)',
        '입금연도', '정액 기술료(원)', '경상 기술료(원)', '지분(주식)현금화(원)',
        '기술료 수입합계(원)', '연구과제명', '연구개발비 재원기관', '연구지원 사업명', '총 연구비(원)'
    ]
    sub_headers = [
        'SEQ','CONT_MNG_NO','CONT_DATE','CONT_CNPR_NM','IDN_FIRM_NM','IDN_FIRM_BRN',
        'IDN_FIRM_AGC_TYP_CD','IDN_FIRM_BTP_TYP_CD','DMT_EXT_DVS','IDN_FIRM_NTN_CD',
        'IDN_FIRM_AEA_CD','TECH_NM','IVR_NM','TECH_TYP_CD','ITL_PRGT_TRANSR_TECH_NUM',
        'ITL_PRGT_NO','TECH_SPHE_CD','TECH_CLS_CD','TR_TYP_CD','CONT_TE',
        'RSRT_ITEM_CD','RSRT_ITEM_ETC','RSRT_ITEM_DESC','TECHFEE_RCVG_TYP_CD',
        'FIX_AMT_TECHFEE_AMT','CAC_TECHFEE_DESC','ACQS_ST_NUM','ST_PA_AMT','RPM_YR',
        'RPM_FIX_AMT_TECHFEE','RPM_CAC_TECHFEE','RPM_FUND_TECHFEE','RPM_TECHFEE_INCME_AMT',
        'RSCH_SBJT_NM','RSCH_DEVFEE_FNRS_AGC_NM','RSCH_SPPT_BIZ_NM','RSRCCT_TOT_AMT'
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    for ci, h in enumerate(sub_headers, 1):
        ws.cell(2, ci, h)

    for idx, row in enumerate(filtered, 1):
        r = idx + 2
        is_foreign = str(row[6]).strip() in ('2', '국외')
        country = str(row[7]) if row[7] else ('대한민국' if not is_foreign else '')
        region  = '해외' if is_foreign else get_region_name(row[8])

        i_fix = safe_int(row[83])
        i_cac = safe_int(row[82])
        i_stk = safe_int(row[77])  # 주식 입금
        total  = safe_int(row[76]) or (i_fix + i_cac + i_stk)
        payment_year = get_year(row[73]) or year

        ws.cell(r, 1,  idx)
        ws.cell(r, 2,  str(row[0]) if row[0] else '')          # 계약관리번호=연번
        ws.cell(r, 3,  to_yyyymmdd(row[1]))                     # 계약일 8자리
        ws.cell(r, 4,  '산학협력단')
        ws.cell(r, 5,  str(row[3]) if row[3] else '')
        ws.cell(r, 6,  str(row[9]) if row[9] else '')           # 사업자번호
        ws.cell(r, 7,  ORG_TYPE_MAP.get(row[4], str(row[4]) if row[4] else ''))
        ws.cell(r, 8,  str(row[5]) if row[5] else '')
        ws.cell(r, 9,  '국외' if is_foreign else '국내')
        ws.cell(r, 10, country)
        ws.cell(r, 11, region)
        ws.cell(r, 12, str(row[28]) if row[28] else '')
        ws.cell(r, 13, str(row[29]) if row[29] else '')
        ws.cell(r, 14, get_text(TECH_TYPE_TEXT, row[37]))
        ws.cell(r, 15, safe_int(row[40]) or 1)
        ws.cell(r, 16, str(row[38]) if row[38] else '')
        ws.cell(r, 17, str(row[42]) if row[42] else '')
        ws.cell(r, 18, str(row[43]) if row[43] else '')
        ws.cell(r, 19, get_text(TRADE_TYPE_TEXT, row[44]))
        ws.cell(r, 20, str(row[45]) if row[45] else '')
        ws.cell(r, 21, '없음')   # 제한사항 기본값 (수동 수정 필요)
        ws.cell(r, 22, '')
        ws.cell(r, 23, '')
        ws.cell(r, 24, str(row[75]) if row[75] else '')         # 기술료수취유형
        ws.cell(r, 25, safe_int(row[52]) or 0)                  # 정액 계약금
        ws.cell(r, 26, str(row[51]) if row[51] else '')         # 경상조건
        ws.cell(r, 27, 0)
        ws.cell(r, 28, 0)
        ws.cell(r, 29, str(payment_year))
        ws.cell(r, 30, i_fix)
        ws.cell(r, 31, i_cac)
        ws.cell(r, 32, i_stk)
        ws.cell(r, 33, total)
        ws.cell(r, 34, str(row[56]) if row[56] else '')
        ws.cell(r, 35, str(row[57]) if row[57] else '')
        ws.cell(r, 36, str(row[59]) if row[59] else '')
        ws.cell(r, 37, safe_int(row[60]) * 1000 if row[60] else 0)  # 천원→원 변환


    print("\n📊 검토 결과 분석 중...")
    write_feedback_sheet(wb, filtered, year, '기술이전 실적 세부 현황', 0)

    wb.save(output_path)
    print(f"\n✅ 저장 완료: {output_path}")
    print(f"   총 {len(filtered)}건")
    print(f"\n⚠️  수동 확인 필요: 제한사항(V열) — 기본값 '없음'으로 입력됨, 개별 확인 후 수정 필요")

if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--master',  required=True)
    p.add_argument('--year',    type=int, required=True)
    p.add_argument('--output',  required=True)
    args = p.parse_args()
    run(args.master, args.year, args.output)
