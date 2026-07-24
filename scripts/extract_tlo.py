"""
TLO혁신형 기술경영촉진사업 기술이전 성과현황 자동 생성
대상 시트: sheet1. 기술이전 성과현황 (13컬럼)

컬럼:
 1.순번  2.주관기관명  3.기술제공기관명(연구자명)  4.기술도입기업명(소재지)
 5.기술명  6.계약일자(YYYYMMDD)  7.총기술료계약액(백만원)  8.입금연도
 9.당해연도기술료입금액(백만원)  10.중대형여부(Y/N)
11.성과구분(①②, 수동입력)  12.계약형태(①정액②경상③노하우)  13.구분
"""
import argparse, sys, os, datetime
sys.path.insert(0, os.path.dirname(__file__))
from common import *
from feedback import write_feedback_sheet
from openpyxl import load_workbook, Workbook


def get_contract_type(row):
    trade     = str(row[44]).strip() if row[44] is not None else ""
    has_fixed = safe_int(row[52]) > 0
    has_cac   = bool(str(row[51]).strip()) if row[51] else False
    if trade in ("3", "8", "노하우", "단위연구노하우"):
        return "③"
    if has_fixed and has_cac:
        return "①+②"
    if has_fixed:
        return "①"
    if has_cac:
        return "②"
    return "①"


def aggregate_by_contract(rows, year):
    contracts = {}
    for row in rows:
        seq = str(row[0]) if row[0] else None
        if not seq:
            continue
        payment = safe_int(row[76])
        p_year  = get_year(row[73])
        if seq not in contracts:
            contracts[seq] = {"row": row, "total_pay": 0, "pay_years": set()}
        contracts[seq]["total_pay"] += payment
        if p_year:
            contracts[seq]["pay_years"].add(p_year)

    result = []
    for seq, info in contracts.items():
        r      = info["row"]
        pay    = info["total_pay"]
        years  = sorted(info["pay_years"])
        pay_yr = year if year in years else (years[-1] if years else year)
        result.append((r, pay, pay_yr))

    result.sort(key=lambda x: x[0][1]
                if isinstance(x[0][1], (datetime.datetime, datetime.date))
                else datetime.datetime(1900, 1, 1))
    return result


def run(master_path, year, output_path,
        org_name="부산대학교 산학협력단", period_str=None, template_path=None):

    print(f"📂 DB 로딩: {master_path}")
    all_rows = load_master_db(master_path)

    filtered = filter_by_year(all_rows, year, mode="contract")
    print(f"  → {year}년 계약 행: {len(filtered)}건")

    filtered = [r for r in filtered if get_bridge_type(r[37], r[44]) is not None]
    print(f"  → 기술자문 제외 후: {len(filtered)}건")

    data = aggregate_by_contract(filtered, year)
    print(f"  → 계약 집계 후: {len(data)}건")

    TARGET = "sheet1. 기술이전 성과현황"
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb[TARGET] if TARGET in wb.sheetnames else wb.active
        if ws.max_row >= 4:
            ws.delete_rows(4, ws.max_row - 3)
    else:
        wb  = Workbook()
        ws  = wb.active
        ws.title = TARGET
        ws.cell(1, 1, "SHEET2. 기술이전 실적현황")
        headers = [
            "순번", "주관기관명", "기술제공기관명(연구자명)",
            "기술도입기업명(소재지)\n*시도기관", "기술명",
            "계약일자\n(8자리, 연도+월+일)",
            "총 기술료 계약액\n(백만원)", "입금연도",
            "당해연도 기술료 입금액\n(백만원)",
            "중대형 기술이전 여부\n(1억원이상 기술료 입금)\n(Yes 또는 No)",
            "성과 구분\n(①기술사업화 프로젝트\n②기술경영촉진)",
            "계약형태 구분\n(①정액 ②경상 ③노하우 )", "구분",
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(3, ci, h)

    DATA_START = 4
    total_payment = 0
    large_count   = 0

    for idx, (row, pay_amt, pay_year) in enumerate(data, 1):
        r = DATA_START + idx - 1

        researcher   = str(row[29]).strip() if row[29] else ""
        provider     = f"부산대학교({researcher})" if researcher else "부산대학교산학협력단"

        company      = str(row[3]).strip() if row[3] else ""
        is_foreign   = str(row[6]).strip() in ("2", "국외")
        region       = "해외" if is_foreign else get_region_name(row[8])
        company_full = f"{company}({region})" if region else company

        contract_won = safe_int(row[52]) + safe_int(row[50])
        contract_m   = round(contract_won / 1_000_000, 1) if contract_won else None
        pay_m        = round(pay_amt / 1_000_000, 1) if pay_amt else None

        is_large = "Y" if pay_amt >= 100_000_000 else "N"
        if is_large == "Y":
            large_count += 1
        total_payment += pay_amt

        ws.cell(r,  1, idx)
        ws.cell(r,  2, org_name)
        ws.cell(r,  3, provider)
        ws.cell(r,  4, company_full)
        ws.cell(r,  5, str(row[28]) if row[28] else "")
        ws.cell(r,  6, to_yyyymmdd(row[1]))
        ws.cell(r,  7, contract_m)
        ws.cell(r,  8, str(pay_year) if pay_year else "")
        ws.cell(r,  9, pay_m)
        ws.cell(r, 10, is_large)
        ws.cell(r, 11, "")
        ws.cell(r, 12, get_contract_type(row))
        ws.cell(r, 13, "")

    raw_rows = [row for row, _, _ in data]
    print("\n📊 검토 결과 분석 중...")
    write_feedback_sheet(wb, raw_rows, year, "TLO혁신형", total_payment)

    wb.save(output_path)
    pay_m_total = round(total_payment / 1_000_000, 1)
    print(f"\n✅ 저장 완료: {output_path}")
    print(f"   총 {len(data)}건 | 입금액 합계: {pay_m_total}백만원 | 중대형: {large_count}건")
    print(f"\n⚠️  수동 입력 필요: K열(성과구분) — ①기술사업화 프로젝트 또는 ②기술경영촉진")
    print(f"📋 검토 결과 시트에서 이상 데이터 확인하세요.")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--master",   required=True)
    p.add_argument("--year",     type=int, required=True)
    p.add_argument("--output",   required=True)
    p.add_argument("--org",      default="부산대학교 산학협력단")
    p.add_argument("--period",   default=None)
    p.add_argument("--template", default=None)
    args = p.parse_args()
    run(args.master, args.year, args.output, args.org, args.period, args.template)
