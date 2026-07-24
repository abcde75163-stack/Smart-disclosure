"""
기술이전 추출 스크립트 공통 유틸리티
모든 extract_*.py 에서 import해서 사용
"""
import re
import datetime

# ── 지역코드 매핑 ──────────────────────────────────────────
REGION_MAP = {
    '02': '서울특별시', '2': '서울특별시',
    '031': '경기도',       '31': '경기도',
    '032': '인천광역시',   '32': '인천광역시',
    '033': '강원도',       '33': '강원도',
    '041': '충청남도',     '41': '충청남도',
    '042': '대전광역시',   '42': '대전광역시',
    '043': '충청북도',     '43': '충청북도',
    '044': '세종특별자치시','44': '세종특별자치시',
    '051': '부산광역시',   '51': '부산광역시',
    '052': '울산광역시',   '52': '울산광역시',
    '053': '대구광역시',   '53': '대구광역시',
    '054': '경상북도',     '54': '경상북도',
    '055': '경상남도',     '55': '경상남도',
    '056': '제주특별자치도','56': '제주특별자치도',
    '061': '전라남도',     '61': '전라남도',
    '062': '광주광역시',   '62': '광주광역시',
    '063': '전라북도',     '63': '전라북도',
    '064': '제주특별자치도','64': '제주특별자치도',
}

# ── 기관유형 매핑 ──────────────────────────────────────────
ORG_TYPE_MAP = {
    1: '대기업', '1': '대기업',
    2: '벤처중소기업', '2': '벤처중소기업',
    3: '일반중소기업', '3': '일반중소기업',
    4: '개인', '4': '개인',
    5: '국공립대학', '5': '국공립대학',
    6: '사립대학', '6': '사립대학',
    7: '국공립시험연구기관', '7': '국공립시험연구기관',
    8: '정부출연연구기관', '8': '정부출연연구기관',
    9: '특정연구기관', '9': '특정연구기관',
    10: '전문생산기술연구소', '10': '전문생산기술연구소',
    11: '기술거래기관', '11': '기술거래기관',
    12: '기타비영리법인및단체', '12': '기타비영리법인및단체',
    13: '해외', '13': '해외',
    14: '기타정부산하기관', '14': '기타정부산하기관',
    15: '기타', '15': '기타',
}

# ── 거래유형 매핑 ──────────────────────────────────────────
TRADE_TYPE_MAP = {
    0: '매매', '0': '매매', '양도(매매)': '매매', '매매': '매매',
    '일부양도': '매매', '0(일부양도)': '매매',
    1: '전용실시', '1': '전용실시', '전용실시': '전용실시',
    2: '통상실시', '2': '통상실시', '통상실시': '통상실시',
    '독점적통상실시': '통상실시', '통상실시(독점)': '통상실시',
    3: '노하우', '3': '노하우', '노하우': '노하우',
    4: '기술제휴', '4': '기술제휴',
    5: 'OEM/ODM', '5': 'OEM/ODM',
    6: 'M&A', '6': 'M&A',
    7: '기타', '7': '기타', '기타': '기타',
    8: '단위연구노하우', '8': '단위연구노하우', '단위연구노하우': '단위연구노하우',
    9: '기술자문', '9': '기술자문', '기술자문': '기술자문',
    '저작권': '저작권',
}

# ── BRIDGE 기술이전유형 매핑 ──────────────────────────────
BRIDGE_EXCLUDE = {9, '9', '기술자문'}

def get_bridge_type(tech_type_raw, trade_type_raw):
    """기술유형 + 거래유형 → BRIDGE 3.0 기술이전유형 (None이면 제외 대상)"""
    tech  = str(tech_type_raw).strip()  if tech_type_raw  is not None else ''
    trade = str(trade_type_raw).strip() if trade_type_raw is not None else ''

    # 기술자문 제외
    if tech in ('9', '기술자문') or trade in ('9', '기술자문'):
        return None
    # 저작권
    if tech in ('10', '저작권') or trade in ('저작권',):
        return '저작권이전'
    # 노하우
    if tech in ('3', '6', '8', '노하우', '단위연구노하우', '정보 및 노하우') or \
       trade in ('3', '8', '노하우', '단위연구노하우'):
        return '노하우이전'
    # 양도/매매
    if trade in ('0', '양도(매매)', '매매', '일부양도', '0(일부양도)'):
        return '산업재산권(상표제외)양도'
    # 전용실시
    if trade in ('1', '전용실시'):
        return '실시권허여-전용'
    # 통상실시
    if trade in ('2', '통상실시', '독점적통상실시', '통상실시(독점)'):
        return '실시권허여-통상'
    # 숫자 코드로 fallback
    try:
        t = int(trade)
        mapping = {0: '산업재산권(상표제외)양도', 1: '실시권허여-전용',
                   2: '실시권허여-통상', 3: '노하우이전', 8: '노하우이전'}
        if t in mapping: return mapping[t]
        if t == 9: return None
    except:
        pass
    # 기술유형 기반 fallback
    if tech in ('1', '특허', '2', '실용신안', '3', '디자인/의장', '4', '상표', '5', '소프트웨어'):
        return '산업재산권(상표제외)양도'
    if tech in ('6', '노하우', '8', '단위연구노하우'):
        return '노하우이전'
    return '기타'

# ── 지역코드 → 지역명 ─────────────────────────────────────
def get_region_name(code_raw, country_code=None):
    if country_code is not None and str(country_code).strip() in ('2', '국외'):
        return '해외'
    if not code_raw:
        return ''
    code_str = str(code_raw).strip()
    match = re.match(r'^(\d+)', code_str)
    if match:
        num = match.group(1).lstrip('0') or '0'
        result = REGION_MAP.get(match.group(1)) or REGION_MAP.get(num, '')
        return result or code_str
    return code_str

# ── 날짜 유틸 ─────────────────────────────────────────────
def to_yyyymmdd(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime('%Y%m%d')
    if d is None: return ''
    s = str(d).strip()
    if re.match(r'^\d{8}$', s): return s
    return s

def to_yyyymm(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime('%Y.%m')
    return ''

def get_year(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.year
    try:
        s = str(d).strip()
        if len(s) >= 8: return int(s[:4])
    except:
        pass
    return None

# ── 금액 유틸 ─────────────────────────────────────────────
def safe_int(v, default=0):
    try:
        if v is None: return default
        return int(float(str(v)))
    except:
        return default

def safe_num(v):
    try:
        if v is None: return None
        f = float(str(v))
        return int(f) if f == int(f) else f
    except:
        return None

def won_to_million(v):
    """원 → 백만원 (반올림 정수)"""
    n = safe_int(v)
    return round(n / 1_000_000, 2) if n else None

def won_to_thousand(v):
    """원 → 천원"""
    n = safe_int(v)
    return n // 1_000 if n else None

# ── DB 로딩 ───────────────────────────────────────────────
def load_master_db(path):
    """마스터 DB 전체 행 로드 (헤더 제외)"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["내역"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue   # 헤더 스킵
        rows.append(row)
    wb.close()
    return rows

def filter_by_year(rows, year, contract_col=1, payment_col=73, mode='both'):
    """
    mode='both'     : 계약일 OR 입금일 중 하나라도 year인 행 (BRIDGE 방식)
    mode='contract' : 계약일이 year인 행만
    mode='payment'  : 입금일이 year인 행만
    """
    result = []
    for row in rows:
        c_year = get_year(row[contract_col])
        p_year = get_year(row[payment_col])
        if mode == 'both':
            if c_year == year or p_year == year:
                result.append(row)
        elif mode == 'contract':
            if c_year == year:
                result.append(row)
        elif mode == 'payment':
            if p_year == year:
                result.append(row)
    return result
