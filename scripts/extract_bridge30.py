"""
BRIDGE 3.0 사업 핵심성과지표 — (3)(증빙) 기술이전 세부 목록 자동 생성
+ 검토 결과 시트 자동 추가

사용법:
  python3 extract_bridge30.py \
    --master "기술이전총정리.xlsx" \
    --template "BRIDGE3.0_원본.xlsx"  [선택] \
    --year 2025 \
    --output "BRIDGE3.0_결과.xlsx"
"""
import argparse, datetime, sys, os
sys.path.insert(0, os.path.dirname(__file__))
from common import *
from feedback import write_feedback_sheet
from openpyxl import load_workbook, Workbook
from collections import defaultdict

def run(master_path, year, output_path, template_path=None):
    print(f"📂 마스터 DB 로딩: {master_path}")
    all_rows = load_master_db(master_path)

    # 계약별 누적 경상기술료 (누적중대형 판단용)
    contract_cumul = defaultdict(int)
    for row in all_rows:
        if row[0]:
            contract_cumul[str(row[0])] += safe_int(row[82])

    # 필터: 계약일 OR 입금일이 year
    filtered = filter_by_year(all_rows, year, mode='both')
    print(f"  → {year}년 해당 행: {len(filtered)}건")

    # 기술자문 제외 + BRIDGE 유형 결정
    data = []
    for row in filtered:
        btype = get_bridge_type(row[37], row[44])
        if btype is None: continue
        data.append((row, btype))

    # 계약일 오름차순 정렬
    data.sort(key=lambda x: x[0][1]
              if isinstance(x[0][1], (datetime.datetime, datetime.date))
              else datetime.datetime(1900, 1, 1))
    print(f"  → 기술자문 제외 후: {len(data)}건")

    # 워크북 준비
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb["(3)(증빙) 기술이전 세부 목록"]
        data_start = 6
        last = ws.max_row
        if last > data_start:
            ws.delete_rows(data_start, last - data_start)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "(3)(증빙) 기술이전 세부 목록"
        ws.cell(2, 1, f'기술이전 실적 세부 목록({year}년)')
        ws.cell(2, 10, '해당기간\n기술이전 건수')
        ws.cell(2, 12, '해당기간\n기술이전 수입료')
        ws.cell(2, 14, '해당기간\n중·대형 기술이전 건수')
        headers = ['연번','계약일자','입금일자','계약명(기술명)',
                   '이전업체명','소재지역(광역)','기술이전유형',
                   '정액기술료\n(계약조건)','경상기술료\n(계약조건)',
                   '정액기술료\n(수입료)','경상기술료\n(수입료)','총수입',
                   '규모구분','AI기술여부','AI기술분류']
        for ci, h in enumerate(headers, 1):
            ws.cell(4, ci, h)
        data_start = 5

    # ── 데이터 채우기 ──────────────────────────────────────
    total_fix = total_cac = total_income = medium_large = 0

    for idx, (row, btype) in enumerate(data, 1):
        r = data_start + idx - 1

        is_foreign = str(row[6]).strip() in ('2', '국외')
        region = '해외' if is_foreign else get_region_name(row[8])

        i_fix = safe_int(row[83])
        i_cac = safe_int(row[82])
        i_tot = safe_int(row[76]) or (i_fix + i_cac)

        total_fix    += i_fix
        total_cac    += i_cac
        total_income += i_tot

        contract_year  = get_year(row[1]) or 0
        fixed_contract = safe_int(row[52])
        cumul          = contract_cumul.get(str(row[0]), 0)

        gyumo = None
        if contract_year == year and fixed_contract >= 100_000_000:
            gyumo = '당해중대형'; medium_large += 1
        elif contract_year < year and cumul >= 100_000_000:
            gyumo = '누적중대형'; medium_large += 1

        ws.cell(r, 1,  idx)
        ws.cell(r, 2,  row[1])
        ws.cell(r, 3,  row[73])
        ws.cell(r, 4,  str(row[28]) if row[28] else '')
        ws.cell(r, 5,  str(row[3])  if row[3]  else '')
        ws.cell(r, 6,  region)
        ws.cell(r, 7,  btype)
        ws.cell(r, 8,  safe_num(row[52]) if safe_int(row[52]) else None)
        ws.cell(r, 9,  str(row[51]) if row[51] else None)
        ws.cell(r, 10, i_fix or None)
        ws.cell(r, 11, i_cac or None)
        ws.cell(r, 12, i_tot or None)
        ws.cell(r, 13, gyumo)
        ws.cell(r, 14, None)   # AI여부: 수동 입력
        ws.cell(r, 15, None)   # AI분류: 수동 입력
        for col in (2, 3):
            ws.cell(r, col).number_format = 'YYYY-MM-DD'

    # 합계행
    sum_r = data_start + len(data)
    ws.cell(sum_r, 9,  '합계')
    ws.cell(sum_r, 10, total_fix)
    ws.cell(sum_r, 11, total_cac)
    ws.cell(sum_r, 12, total_income)

    # 헤더 요약
    ws.cell(2, 11, f'{len(data)}건')
    ws.cell(2, 13, total_income)
    ws.cell(2, 15, f'{medium_large}건')

    # ── 검토 결과 시트 추가 ───────────────────────────────
    print("\n📊 검토 결과 분석 중...")
    raw_rows = [row for row, _ in data]
    write_feedback_sheet(wb, raw_rows, year, 'BRIDGE 3.0', total_income)

    wb.save(output_path)
    print(f"\n✅ 저장 완료: {output_path}")
    print(f"   총 {len(data)}건 | 총수입: {total_income:,}원 | 중대형: {medium_large}건")
    print(f"\n⚠️  수동 입력 필요: AI 기술 여부 (N열), AI 기술분류 (O열)")
    print(f"📋 '검토 결과' 시트에서 이상 데이터 및 확인 필요 항목을 확인하세요.")

if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--master',   required=True)
    p.add_argument('--year',     type=int, required=True)
    p.add_argument('--output',   required=True)
    p.add_argument('--template', default=None)
    args = p.parse_args()
    run(args.master, args.year, args.output, args.template)
