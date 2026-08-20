"""
양식 기반 AI 자동 추출기 — 모든 양식에 대해 작동

핵심 원칙:
  양식 파일 안에 있는 가이드 지침을 AI가 직접 읽고 그에 따라 추출한다.
  하드코딩된 매핑 없음. 양식 파일이 곧 명세서.
"""
import json, sys, os, re, datetime
sys.path.insert(0, os.path.dirname(__file__))
from common import *
from ai_client import create_json_response
from feedback import write_feedback_sheet
from analyze_template import analyze
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_notes_filters(notes: str) -> dict:
    """
    추가 요청사항 텍스트에서 날짜 범위·필터 기준·연구자 목록을 파싱.
    반환: {"date_start": "YYYYMMDD"|None, "date_end": "YYYYMMDD"|None,
           "date_mode": "contract"|"payment"|"both",
           "researchers": [이름,...] | None}
    """
    result = {"date_start": None, "date_end": None,
              "date_mode": "both", "researchers": None}
    if not notes:
        return result

    # 날짜 범위: 20260101~20260801 / 20260101-20260801 / 20260101 ~ 20260801
    m = re.search(r'(20\d{6})\s*[~\-~]\s*(20\d{6})', notes)
    if m:
        result["date_start"] = m.group(1)
        result["date_end"]   = m.group(2)

    # 필터 기준
    if "계약일" in notes:
        result["date_mode"] = "contract"
    elif "입금일" in notes:
        result["date_mode"] = "payment"

    # 연구자 목록: "대상 연구자: 홍길동, 김철수" 또는 "연구자: ..."
    rm = re.search(r'(?:대상\s*)?연구자\s*[:：]\s*(.+)', notes)
    if rm:
        names_raw = rm.group(1).strip()
        names = [n.strip() for n in re.split(r'[,，/\n]', names_raw) if n.strip()]
        if names:
            result["researchers"] = names

    return result

# ── 매핑 캐시 ──────────────────────────────────────────────────────
_CACHE_PATH = os.path.join(os.path.dirname(__file__), "..", "references", "mapping_cache.json")

def _normalize_key(filename: str) -> str:
    """파일명 → 캐시 조회용 정규화 키."""
    name = os.path.splitext(os.path.basename(filename))[0].lower()
    for w in ["자동추출", "수정안내", "복사본", "최종", "수정", "제출용", "붙임", "별첨"]:
        name = name.replace(w, "")
    name = re.sub(r'20\d{2}', '', name)          # 연도 제거
    name = re.sub(r'\d{4,}', '', name)            # 긴 숫자 제거
    name = re.sub(r'[\s_\-\(\)\[\]\.]+', ' ', name).strip()
    return name

def load_cache() -> dict:
    try:
        with open(_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def find_cached_mapping(template_filename: str, cache: dict) -> tuple:
    """파일명으로 캐시에서 가장 유사한 매핑 탐색. (key, mapping) 또는 (None, None)."""
    key = _normalize_key(template_filename)
    best_key, best_mapping, best_score = None, None, 0
    for cache_key, entry in cache.items():
        if cache_key.startswith("_"):
            continue
        mapping = entry.get("mapping") if isinstance(entry, dict) else None
        if not mapping:
            continue
        # 부분 문자열 매칭 점수
        score = 0
        for part in cache_key.split():
            if part and part in key:
                score += len(part)
        for part in key.split():
            if part and part in cache_key:
                score += len(part)
        if score > best_score:
            best_score, best_key, best_mapping = score, cache_key, mapping
    if best_score >= 4:   # 최소 4자 이상 겹쳐야 매칭으로 인정
        return best_key, best_mapping
    return None, None

def save_mapping_to_cache(template_filename: str, mapping: dict, cache: dict,
                          example_rows: list = None) -> dict:
    """추출 성공 후 매핑(+수동 검증 예시 행)을 캐시에 추가."""
    key = _normalize_key(template_filename)
    entry = cache.get(key, {})
    entry.update({
        "saved_at": datetime.datetime.now().strftime("%Y-%m-%d"),
        "template_hint": template_filename,
        "mapping": mapping,
    })
    if example_rows is not None:
        entry["example_rows"] = example_rows   # 수동 검증된 올바른 출력 행
    cache[key] = entry
    return cache


def extract_example_rows(corrected_file_bytes: bytes, target_sheet: str = None,
                          data_start: int = 4, n_rows: int = 5) -> list:
    """수정된 결과 파일에서 헤더 + 데이터 예시 행 추출."""
    import io
    wb = load_workbook(io.BytesIO(corrected_file_bytes), data_only=True)
    ws = None
    if target_sheet and target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
    else:
        ws = wb.active

    header_row = data_start - 1
    rows_out = []
    for r in range(header_row, min(header_row + n_rows + 1, (ws.max_row or 0) + 1)):
        cells = {}
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                cells[str(c)] = str(v).strip()
        if cells:
            rows_out.append({"row": r, "cells": cells, "is_header": r == header_row})
    return rows_out


def format_analysis_for_prompt(template_analysis, title="파일 내용"):
    """analyze_template 결과를 AI 프롬프트에 넣기 쉬운 텍스트로 변환."""
    text_parts = [f"\n## {title}"]
    for sheet in template_analysis.get("sheets", []):
        rows = sheet.get("rows", [])
        text_parts.append(f"\n### 시트: {sheet['name']} (최대행:{sheet.get('max_row','?')}, 최대열:{sheet.get('max_col','?')})")

        header_row_num = None
        max_cells = 0
        for row_info in rows[:10]:
            if len(row_info["cells"]) > max_cells:
                max_cells = len(row_info["cells"])
                header_row_num = row_info["row"]

        for row_info in rows:
            cells_str = ", ".join(f"[열{k}]{v}" for k, v in row_info["cells"].items())
            if row_info["row"] == header_row_num:
                text_parts.append(f"  ★헤더행{row_info['row']}: {cells_str}")
            elif row_info["row"] > (header_row_num or 0):
                text_parts.append(f"  데이터행{row_info['row']}: {cells_str}")
            else:
                text_parts.append(f"  행{row_info['row']}: {cells_str}")
        if sheet.get("merged_cells"):
            text_parts.append(f"  병합셀: {', '.join(sheet['merged_cells'][:40])}")
        if sheet.get("hidden_rows"):
            text_parts.append(f"  숨김행: {sheet['hidden_rows'][:40]}")
        if sheet.get("hidden_cols"):
            text_parts.append(f"  숨김열: {sheet['hidden_cols'][:40]}")
        if sheet.get("formulas"):
            text_parts.append("  수식:")
            for item in sheet["formulas"][:40]:
                text_parts.append(f"    {item['cell']}: {item['formula']}")
        if sheet.get("comments"):
            text_parts.append("  셀 주석/메모:")
            for item in sheet["comments"][:40]:
                text_parts.append(f"    {item['cell']}: {item['text']}")
        if sheet.get("style_hints"):
            text_parts.append("  서식 힌트(색/굵게/줄바꿈은 지침·헤더 구분 단서):")
            for item in sheet["style_hints"][:60]:
                text_parts.append(
                    f"    {item['cell']}: {item['value']} | fill={item.get('fill')} | font={item.get('font')}"
                )
        if sheet.get("truncated"):
            text_parts.append("  (이하 행 생략)")
    return "\n".join(text_parts)


def _norm_match_value(value):
    """명단 파일과 DB 값을 비교하기 위한 느슨한 문자열 정규화."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return re.sub(r"[\s\-_/]", "", s)


def extract_reference_values(reference_path, reference_filter):
    """요청파일의 특정 열에서 대상자/대상값 목록을 추출."""
    sheet_name = reference_filter.get("sheet")
    col = int(reference_filter.get("column") or reference_filter.get("col") or 0)
    start_row = int(reference_filter.get("data_start_row") or 2)
    if col < 1:
        return []

    wb = load_workbook(reference_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    values = []
    seen = set()
    for r in range(start_row, (ws.max_row or 0) + 1):
        raw = ws.cell(r, col).value
        normalized = _norm_match_value(raw)
        if normalized and normalized not in seen:
            values.append(normalized)
            seen.add(normalized)
    wb.close()
    return values


def infer_reference_filters(reference_paths, notes=""):
    """AI가 필터를 주지 못했을 때 요청파일 헤더에서 흔한 대상 열을 추정."""
    if not reference_paths:
        return []
    if isinstance(reference_paths, str):
        reference_paths = [reference_paths]

    wants_reference_filter = any(k in (notes or "") for k in ["동일", "대상", "명단", "교직원", "교번", "연구자", "교수"])
    if not wants_reference_filter:
        return []

    for file_index, reference_path in enumerate(reference_paths, 1):
        wb = load_workbook(reference_path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            max_scan_row = min(ws.max_row or 0, 10)
            for r in range(1, max_scan_row + 1):
                for c in range(1, (ws.max_column or 0) + 1):
                    header = ws.cell(r, c).value
                    if not header:
                        continue
                    header_text = str(header).replace("\n", " ").strip()
                    if any(k in header_text for k in ["교직원", "교번", "직원번호", "사번"]):
                        wb.close()
                        return [{
                            "source": "request_file",
                            "source_file_index": file_index,
                            "sheet": ws.title,
                            "column": c,
                            "data_start_row": r + 1,
                            "db_indices": [30, 34],
                            "match_type": "contains",
                            "label": header_text,
                        }]
                    if any(k in header_text for k in ["연구자", "교수", "성명", "이름", "발명자"]):
                        wb.close()
                        return [{
                            "source": "request_file",
                            "source_file_index": file_index,
                            "sheet": ws.title,
                            "column": c,
                            "data_start_row": r + 1,
                            "db_indices": [29, 33],
                            "match_type": "contains",
                            "label": header_text,
                        }]
        wb.close()
    return []


def apply_reference_filters(rows, reference_paths, reference_filters):
    """요청파일에서 추출한 대상값으로 마스터 DB 행을 필터링."""
    if not reference_paths or not reference_filters:
        return rows
    if isinstance(reference_paths, str):
        reference_paths = [reference_paths]

    filtered = rows
    for reference_filter in reference_filters:
        if reference_filter.get("source") not in (None, "", "request_file", "supplemental_file"):
            continue
        source_index = int(reference_filter.get("source_file_index") or 1)
        if source_index < 1 or source_index > len(reference_paths):
            print(f"  ⚠️ 요청파일 필터 source_file_index 범위 오류: {source_index}")
            continue

        values = set(extract_reference_values(reference_paths[source_index - 1], reference_filter))
        db_indices = []
        for db_idx in reference_filter.get("db_indices") or []:
            try:
                db_indices.append(int(db_idx))
            except Exception:
                pass
        if not values or not db_indices:
            print(f"  ⚠️ 요청파일 필터 '{reference_filter.get('label', '')}' 값 또는 DB 컬럼 없음")
            continue

        before = len(filtered)
        match_type = reference_filter.get("match_type", "exact")

        def row_matches(row):
            for db_idx in db_indices:
                if db_idx >= len(row):
                    continue
                cell_value = _norm_match_value(row[db_idx])
                if not cell_value:
                    continue
                if match_type == "exact" and cell_value in values:
                    return True
                if match_type != "exact" and any(v in cell_value or cell_value in v for v in values):
                    return True
            return False

        filtered = [row for row in filtered if row_matches(row)]
        print(
            f"  → 요청파일 대상 필터 '{reference_filter.get('label', '')}': "
            f"{len(values)}개 값 / {before}건 → {len(filtered)}건"
        )
    return filtered


def _db_filter_values(db_filter):
    """db_filters의 value/values를 리스트로 정규화."""
    if "values" in db_filter and db_filter.get("values") is not None:
        raw = db_filter.get("values")
    elif "value" in db_filter:
        raw = [db_filter.get("value")]
    else:
        raw = []
    if not isinstance(raw, list):
        raw = [raw]
    return [v for v in raw if v is not None and str(v).strip() != ""]


def _compare_as_number(value):
    try:
        if value is None or str(value).strip() == "":
            return None
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def _compare_as_date_text(value):
    text = to_yyyymmdd(value)
    if not text:
        return ""
    return re.sub(r"\D", "", str(text))[:8]


def _cell_matches_filter(cell_value, db_filter):
    """단일 셀 값이 db_filter 조건을 만족하는지 판단."""
    operator = str(db_filter.get("operator") or "eq").lower()
    values = _db_filter_values(db_filter)
    transform = str(db_filter.get("value_type") or db_filter.get("type") or "auto").lower()

    if transform == "number" or operator in ("gt", "gte", "lt", "lte"):
        cell_num = _compare_as_number(cell_value)
        nums = [_compare_as_number(v) for v in values]
        nums = [n for n in nums if n is not None]
        if cell_num is None:
            return False
        if operator == "gt":
            return bool(nums) and cell_num > nums[0]
        if operator == "gte":
            return bool(nums) and cell_num >= nums[0]
        if operator == "lt":
            return bool(nums) and cell_num < nums[0]
        if operator == "lte":
            return bool(nums) and cell_num <= nums[0]
        if operator == "between":
            return len(nums) >= 2 and min(nums[0], nums[1]) <= cell_num <= max(nums[0], nums[1])
        if operator in ("eq", "in"):
            return cell_num in nums
        if operator in ("neq", "not_in"):
            return cell_num not in nums

    if transform in ("date", "yyyymmdd") or operator in ("date_between", "year_eq"):
        cell_date = _compare_as_date_text(cell_value)
        date_values = [_compare_as_date_text(v) for v in values]
        date_values = [v for v in date_values if v]
        if not cell_date:
            return False
        if operator in ("between", "date_between"):
            return len(date_values) >= 2 and min(date_values[0], date_values[1]) <= cell_date <= max(date_values[0], date_values[1])
        if operator == "year_eq":
            years = [str(v)[:4] for v in values if str(v).strip()]
            return cell_date[:4] in years
        if operator in ("eq", "in"):
            return cell_date in date_values
        if operator in ("neq", "not_in"):
            return cell_date not in date_values

    cell_text = str(cell_value).strip() if cell_value is not None else ""
    cell_norm = _norm_match_value(cell_value)
    value_texts = [str(v).strip() for v in values]
    value_norms = [_norm_match_value(v) for v in values]

    if operator in ("eq", "equals"):
        return any(cell_text == v or cell_norm == n for v, n in zip(value_texts, value_norms))
    if operator in ("neq", "not_equals"):
        return not any(cell_text == v or cell_norm == n for v, n in zip(value_texts, value_norms))
    if operator == "in":
        return any(cell_text == v or cell_norm == n for v, n in zip(value_texts, value_norms))
    if operator == "not_in":
        return not any(cell_text == v or cell_norm == n for v, n in zip(value_texts, value_norms))
    if operator == "contains":
        return any(v and (v in cell_text or n in cell_norm) for v, n in zip(value_texts, value_norms))
    if operator == "not_contains":
        return not any(v and (v in cell_text or n in cell_norm) for v, n in zip(value_texts, value_norms))
    if operator in ("is_empty", "empty"):
        return cell_text == ""
    if operator in ("not_empty", "is_not_empty"):
        return cell_text != ""

    return False


def apply_db_filters(rows, db_filters):
    """AI가 생성한 일반 마스터 DB 필터를 적용."""
    if not db_filters:
        return rows

    filtered = rows
    for db_filter in db_filters:
        try:
            db_index = int(db_filter.get("db_index"))
        except Exception:
            print(f"  ⚠️ DB 필터 db_index 오류: {db_filter}")
            continue
        if db_index < 0:
            continue

        before = len(filtered)
        label = db_filter.get("label") or f"DB[{db_index}]"
        filtered = [
            row for row in filtered
            if db_index < len(row) and _cell_matches_filter(row[db_index], db_filter)
        ]
        print(f"  → DB 조건 필터 '{label}': {before}건 → {len(filtered)}건")
    return filtered

# ── 지원 transform 함수 ────────────────────────────────────────────
TRANSFORMS = {
    "str":               lambda row, idx: str(row[idx]) if row[idx] is not None else "",
    "date_yyyymmdd":     lambda row, idx: to_yyyymmdd(row[idx]),
    "date_yyyymm":       lambda row, idx: to_yyyymm(row[idx]),
    "date_year":         lambda row, idx: str(get_year(row[idx])) if get_year(row[idx]) else "",
    "date_obj":          lambda row, idx: row[idx],
    "safe_int":          lambda row, idx: safe_int(row[idx]),
    "won_to_million":    lambda row, idx: round(safe_int(row[idx]) / 1_000_000, 1) if safe_int(row[idx]) else None,
    "won_to_thousand":   lambda row, idx: safe_int(row[idx]) // 1_000 if safe_int(row[idx]) else None,
    "region":            lambda row, idx: get_region_name(row[8], row[6]),
    "org_type":          lambda row, idx: ORG_TYPE_MAP.get(row[idx], str(row[idx]) if row[idx] else ""),
    "trade_type":        lambda row, idx: TRADE_TYPE_MAP.get(row[idx], str(row[idx]) if row[idx] else ""),
    "domestic_foreign":  lambda row, idx: "국외" if str(row[6]).strip() in ("2", "국외") else "국내",
    "bridge_type":       lambda row, idx: get_bridge_type(row[37], row[44]) or "",
}


def run_free_format(all_rows, template_path, mapping, year, output_path):
    """
    연구자/교수 명단에서 이름 추출 → 해당 연구자의 기술이전 실적을 자유 양식으로 생성.
    form_type='researcher_list'일 때 호출됨.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── 연구자 명단 읽기 ──────────────────────────────────────────────
    wb_form   = load_workbook(template_path, data_only=True)
    res_sheet = mapping.get("researcher_sheet")
    ws_form   = wb_form[res_sheet] if res_sheet and res_sheet in wb_form.sheetnames else wb_form.active

    res_col   = mapping.get("researcher_col", 2)
    res_start = mapping.get("researcher_data_start", 2)

    researcher_names = []
    for r in range(res_start, (ws_form.max_row or 0) + 1):
        v = ws_form.cell(r, res_col).value
        if v and str(v).strip():
            name = str(v).strip()
            if name and name not in researcher_names:
                researcher_names.append(name)

    print(f"  👥 명단에서 연구자 {len(researcher_names)}명 추출: {researcher_names[:5]}")

    # ── 연구자 이름으로 마스터 DB 필터링 ────────────────────────────
    def researcher_match(row):
        investigator = str(row[29]).strip() if row[29] else ""
        if not investigator:
            return False
        return any(
            (len(name) >= 2) and (name in investigator or investigator in name)
            for name in researcher_names
        )

    matched_all = [r for r in all_rows if researcher_match(r)]

    # 연도 필터 시도 → 없으면 전체 기간
    year_filtered = filter_by_year(matched_all, year, mode="both")
    if year_filtered:
        rows_to_use = year_filtered
        year_label  = f"{year}년"
    else:
        rows_to_use = matched_all
        year_label  = "전체 기간"
        print(f"  ⚠️ {year}년 기준 결과 없음 → 전체 기간으로 확장")

    print(f"  → 필터링 결과: {len(rows_to_use)}건 ({year_label})")

    # ── 계약 단위로 집계 (동일 연번 입금 회차 합산) ─────────────────
    contracts = {}
    for row in rows_to_use:
        seq = str(row[0]) if row[0] else ""
        if not seq:
            continue
        payment = safe_int(row[76])
        if seq not in contracts:
            contracts[seq] = {"row": row, "total_pay": 0}
        contracts[seq]["total_pay"] += payment

    sorted_contracts = sorted(
        contracts.values(),
        key=lambda x: x["row"][1]
        if isinstance(x["row"][1], (datetime.datetime, datetime.date))
        else datetime.datetime(1900, 1, 1),
    )

    # ── 새 워크북 생성 ────────────────────────────────────────────────
    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "기술이전 실적"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_center = Alignment(horizontal="center", vertical="center")
    data_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        ("순번",                   6),
        ("연구자명",               12),
        ("기술명",                 30),
        ("기술도입업체(지역)",     22),
        ("계약일",                 12),
        ("총기술료계약액\n(백만원)", 14),
        (f"{year_label}\n입금액(백만원)", 14),
        ("계약형태",               10),
        ("중대형여부\n(1억 이상)", 10),
    ]

    # 제목 행 (1행)
    title_cell = ws.cell(1, 1, f"기술이전 실적 현황 ({year_label})")
    title_cell.font = Font(bold=True, size=12, color="1F3864")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 헤더 행 (2행)
    ws.row_dimensions[2].height = 36
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    # 데이터 행 (3행~)
    total_income = 0
    for idx, info in enumerate(sorted_contracts, 1):
        row      = info["row"]
        pay_total = info["total_pay"]
        total_income += pay_total

        company    = str(row[3]).strip() if row[3] else ""
        is_foreign = str(row[6]).strip() in ("2", "국외")
        region     = "해외" if is_foreign else get_region_name(row[8])
        company_str = f"{company}({region})" if region else company

        contract_m = round((safe_int(row[52]) + safe_int(row[50])) / 1_000_000, 1)
        pay_m      = round(pay_total / 1_000_000, 1) if pay_total else None

        row_num = idx + 2
        ws.row_dimensions[row_num].height = 20

        data_cells = [
            (idx,                                        data_center),
            (str(row[29]).strip() if row[29] else "",   data_center),
            (str(row[28]).strip() if row[28] else "",   data_left),
            (company_str,                                data_left),
            (to_yyyymmdd(row[1]),                        data_center),
            (contract_m if contract_m else None,         data_center),
            (pay_m,                                      data_center),
            (get_contract_type(row),                     data_center),
            ("Y" if pay_total >= 100_000_000 else "N",  data_center),
        ]
        for c, (val, align) in enumerate(data_cells, 1):
            cell = ws.cell(row_num, c, val)
            cell.alignment = align
            cell.border    = border
            if c in (6, 7) and isinstance(val, float):
                cell.number_format = "0.0"

    # 합계 행
    if sorted_contracts:
        sum_row  = len(sorted_contracts) + 3
        sum_fill = PatternFill("solid", fgColor="D6E4FF")
        sum_font = Font(bold=True)
        ws.row_dimensions[sum_row].height = 22
        for c in range(1, len(headers) + 1):
            cell = ws.cell(sum_row, c)
            cell.fill      = sum_fill
            cell.font      = sum_font
            cell.border    = border
            cell.alignment = data_center
        ws.cell(sum_row, 1, "합계")
        ws.cell(sum_row, 6).number_format = "0.0"
        ws.cell(sum_row, 7).number_format = "0.0"
        ws.cell(sum_row, 6, round(
            sum(safe_int(i["row"][52]) + safe_int(i["row"][50]) for i in sorted_contracts) / 1_000_000, 1
        ))
        ws.cell(sum_row, 7, round(total_income / 1_000_000, 1) if total_income else None)

    wb_out.save(output_path)
    print(f"\n✅ 자유 양식 저장 완료: {output_path}")
    print(f"   연구자 {len(researcher_names)}명 | 실적 {len(sorted_contracts)}건 | 총입금 {total_income:,}원")

    return {
        "count":            len(sorted_contracts),
        "total_income":     total_income,
        "mapping_notes":    f"연구자 명단 {len(researcher_names)}명 기준 자유 양식 추출 ({year_label})",
        "manual_inputs":    [],
        "filter_mode":      year_label,
        "mapping":          mapping,
        "cache_key":        None,
        "form_type":        "researcher_list",
        "researcher_count": len(researcher_names),
    }


def get_mapping_from_openai(template_analysis, db_columns_text, transform_rules_text, notes, year, hint,
                            api_key=None, cached_mapping=None, previous_mapping=None,
                            validation_feedback=None, request_files_context=None,
                            request_file_count=1):
    """
    양식 파일 전체 내용을 AI에 전달하고 컬럼 매핑 JSON을 받아온다.
    cached_mapping: 유사 양식에서 성공한 이전 매핑 (참고용으로 프롬프트에 주입)
    """
    form_full_text = request_files_context or format_analysis_for_prompt(template_analysis, "요청파일 1 내용")

    if cached_mapping:
        cached_section = (
            "이전에 유사한 양식에서 성공적으로 사용된 매핑입니다. "
            "현재 양식의 헤더와 지침을 우선하되, 구조가 비슷하면 참고하세요.\n"
            f"```json\n{json.dumps(cached_mapping, ensure_ascii=False, indent=2)}\n```"
        )
        # 수동 검증된 예시 행이 있으면 추가
        example_rows = None
        if cached_mapping:
            cache = load_cache()
            key = _normalize_key(template_analysis.get("_filename", ""))
            _, found_mapping = find_cached_mapping(template_analysis.get("_filename", ""), cache)
            for ck, ce in cache.items():
                if isinstance(ce, dict) and ce.get("mapping") == cached_mapping:
                    example_rows = ce.get("example_rows")
                    break
        if example_rows:
            rows_text = []
            for row_info in example_rows:
                tag = "헤더" if row_info.get("is_header") else "데이터"
                cells_str = ", ".join(f"[열{k}]{v}" for k, v in row_info["cells"].items())
                rows_text.append(f"  {tag}행{row_info['row']}: {cells_str}")
            cached_section += (
                "\n\n### 수동 검증된 올바른 출력 예시 (이 형식을 정확히 따르세요)\n"
                + "\n".join(rows_text)
            )
    else:
        cached_section = "(없음 — 처음 보는 양식이므로 양식 파일 지침만 따르세요)"

    if previous_mapping and validation_feedback:
        validation_section = (
            "아래는 직전 매핑을 코드가 자동 검증한 결과입니다. "
            "문제를 수정한 새 JSON 매핑을 다시 생성하세요.\n"
            "직전 매핑:\n"
            f"```json\n{json.dumps(previous_mapping, ensure_ascii=False, indent=2)}\n```\n"
            "검증 피드백:\n"
            f"```json\n{json.dumps(validation_feedback, ensure_ascii=False, indent=2)}\n```"
        )
    else:
        validation_section = "(첫 번째 매핑 생성 단계)"

    prompt = f"""당신은 기술이전 데이터 추출 전문가입니다.

아래에 제공된 요청파일들의 전체 내용을 읽고, 사용자의 자연어 요청에 따라 어떤 파일이 최종 작성양식인지, 어떤 파일이 작성요령/추출대상인지 판단한 뒤 컬럼 매핑 JSON을 생성해주세요.

## 중요 원칙
- 마스터 DB 파일은 항상 동일한 고정 스키마입니다. 요청파일들은 마스터 DB가 아니라 작성양식/작성요령/추출대상자 명단입니다.
- 최종 결과를 채워 넣을 작성양식 파일 번호를 `output_file_index`로 지정하세요. 사용자가 자연어로 "요청파일 3 양식"처럼 말하면 그 번호를 따르세요.
- 작성양식 파일 내부의 가이드 지침, 예시 데이터, 컬럼 헤더에 명시된 형식을 최우선으로 따르세요.
- 작성요령 파일이 따로 있으면 인정기준, 제외대상, 기간 기준, 단위 규칙을 매핑에 반영하세요.
- 추출대상자 명단 파일이 따로 있고 사용자가 "동일한 대상", "명단 기준", "교직원번호 기준" 등을 요청하면 `reference_filters`에 대상 필터를 작성하세요.
- 사용자가 자연어로 마스터 DB 조건을 말하면 `db_filters`에 구조화하세요. 예: "입금액 1억원 이상", "부처명이 교육부", "기술자문 제외", "업체명에 바이오 포함", "계약일 20260101~20260630".
- 헤더에 "(백만원)"이라고 쓰여 있으면 백만원 단위로, "(YYYYMMDD)"이면 8자리 날짜로, "(Yes/No)"이면 Y/N으로 변환하세요.
- 예시 행(XX대학교, 홍길동 등)의 형식을 정확히 따르세요. 예) "XX대학교(홍길동)" 형식이면 "부산대학교(연구자명)" 형식으로.
- 작성 지침에 인정기준이 있으면 그에 맞게 필터 조건을 설정하세요.

## 추출 연도
{year}년

## 이전 유사 양식의 성공 매핑 (참고용)
{cached_section}

## 자동 검증 피드백
{validation_section}

## 요청파일 전체 내용
총 요청파일 수: {request_file_count}개
{form_full_text}

## 마스터 DB 컬럼 목록
{db_columns_text}

## 변환 규칙
{transform_rules_text}

## 사용자 추가 요청사항
{notes if notes else "(없음)"}

## 키워드 힌트 (참고용)
{hint if hint else "(없음)"}

## 마스터 DB 주요 컬럼 참조 (db_index 혼동 방지용)
마스터 DB(기술이전총정리.xlsx)는 항상 동일한 구조입니다.
작성양식 파일에 해당 항목이 있을 때만 아래 인덱스를 사용하세요.
- db_index 1  → 계약일 (기술이전계약일)
- db_index 3  → 기술도입업체명 (기관/업체명)
- db_index 8  → 지역코드 (국내지역구분 코드)
- db_index 28 → 기술명
- db_index 29 → 주발명자/연구자명
- db_index 37 → 기술유형
- db_index 44 → 거래유형
- db_index 51 → 경상기술료 조건 텍스트
- db_index 52 → 정액기술료 계약금(원)
- db_index 73 → 입금일
- db_index 76 → 현금입금액(원)
- db_index 82 → 경상기술료 입금액(원)
- db_index 83 → 정액기술료 입금액(원)
작성양식 파일에 없는 항목은 절대 포함하지 마세요.

---

## 특수 케이스: 연구자/교수 명단 파일
양식 파일이 기술이전 성과를 기록하는 출력 양식이 아니라, **연구자·교수·인원 명단**(예: 교수님 명단, BK참여 명단, 담당자 리스트 등)으로 판단되면, 아래 형식으로만 응답하세요:

```json
{{
  "form_type": "researcher_list",
  "researcher_sheet": "연구자명이 있는 시트명",
  "researcher_col": 2,
  "researcher_data_start": 2,
  "notes": "연구자 명단 파일로 감지됨. 해당 연구자들의 기술이전 실적을 자유 양식으로 추출."
}}
```
- `researcher_col`: 연구자/교수명이 있는 열 번호(1-based)
- `researcher_data_start`: 데이터가 시작되는 행 번호(헤더 다음 행)
- 이 경우 columns 배열 불필요

판단 기준: 시트에 BK참여, 교번, 임용일, 과학인등록번호 같은 인사 정보가 있고, 기술이전 실적 기재 열이 없으면 연구자 명단.

---

반드시 아래 JSON 형식으로만 응답하세요. 설명 없이 JSON 코드블록만 출력하세요.

```json
{{
  "form_type": "output_form",
  "output_file_index": 1,
  "target_sheet": "데이터를 채울 시트명",
  "data_start_row": 4,
  "filter_mode": "contract",
  "exclude_consulting": true,
  "group_by_contract": false,
  "reference_filters": [
    {{
      "source": "request_file",
      "source_file_index": 1,
      "sheet": "대상자 명단",
      "column": 2,
      "data_start_row": 2,
      "db_indices": [30, 34],
      "match_type": "contains",
      "label": "교직원번호"
    }}
  ],
  "db_filters": [
    {{
      "db_index": 76,
      "operator": "gte",
      "value": 100000000,
      "value_type": "number",
      "label": "현금입금액 1억원 이상"
    }},
    {{
      "db_index": 57,
      "operator": "contains",
      "value": "교육부",
      "value_type": "text",
      "label": "부처명 교육부 포함"
    }}
  ],
  "confidence": {{
    "overall": 0.86,
    "reason": "헤더, 작성 지침, 예시 행이 대부분 일치하나 성과구분은 DB에 직접 컬럼이 없어 수동 확인 필요"
  }},
  "columns": [
    {{
      "col": 1,
      "type": "sequence",
      "label": "순번",
      "confidence": 0.99,
      "evidence": "헤더가 순번이며 데이터 행 순서값 요구"
    }},
    {{
      "col": 2,
      "type": "db_value",
      "db_index": 3,
      "transform": "str",
      "label": "도입업체명",
      "confidence": 0.92,
      "evidence": "헤더가 도입업체/기업명이며 마스터 DB 3번 컬럼과 의미 일치"
    }},
    {{
      "col": 3,
      "type": "constant",
      "value": "부산대학교산학협력단",
      "label": "기술공급기관"
    }},
    {{
      "col": 4,
      "type": "researcher_concat",
      "label": "기술제공기관명(연구자명)"
    }},
    {{
      "col": 5,
      "type": "company_region",
      "label": "기술도입기업명(소재지)"
    }},
    {{
      "col": 6,
      "type": "sum",
      "db_indices": [82, 83],
      "unit": "million",
      "label": "기술료합계(백만원)"
    }}
  ],
  "manual_inputs": ["성과구분(K열) — ①②직접 선택 필요"],
  "notes": "매핑 특이사항 또는 주의사항"
}}
```

## column type 설명
- "sequence": 행 순번(1,2,3...)
- "db_value": DB의 특정 컬럼값 (db_index + transform 필수)
- "constant": 고정 문자열 (value 필수)
- "sum": 여러 DB 컬럼 합산 (db_indices + unit 필수). ※group_by_contract=true일 때 입금액 합산에는 사용 금지 — payment_amount 사용할 것
- "payment_amount": 집계된 당해연도 입금액(unit 필수). group_by_contract=true일 때 "당해연도 기술료 입금액" 같은 컬럼에 반드시 사용
- "researcher_concat": "부산대학교(연구자명)" 형식 자동 생성
- "company_region": "업체명(지역)" 형식 자동 생성
- "payment_year": 입금일에서 연도 추출
- "large_transfer": 입금액 1억 이상 여부 (Y/N)
- "contract_type": 계약형태 자동 판단 (①정액/②경상/③노하우)

## 중요: 헤더가 있는 모든 열을 빠짐없이 매핑하세요
양식 파일의 데이터 행이 비어있어도 무시하세요. 헤더 행 기준으로 모든 열을 매핑합니다.
예: 헤더에 "기술명"이 있으면 반드시 db_value(db_index=28)로 매핑해야 합니다.

## confidence 작성 규칙
- 전체 매핑 confidence.overall은 0~1 숫자로 작성하세요.
- 각 columns 항목에도 confidence와 evidence를 작성하세요.
- 0.70 미만인 컬럼은 manual_inputs에도 넣고, 왜 불확실한지 notes에 적으세요.
- 양식 지침과 DB 컬럼이 직접 대응되지 않으면 억지 매핑하지 말고 constant 또는 manual_inputs로 분리하세요.

## transform 목록
str, date_yyyymmdd, date_yyyymm, date_year, date_obj, safe_int,
won_to_million, won_to_thousand, region, org_type, trade_type,
domestic_foreign, bridge_type

## filter_mode
- "both": 계약일 또는 입금일이 해당 연도 (BRIDGE 방식)
- "contract": 계약일 기준
- "payment": 입금일 기준

## group_by_contract
- true: 동일 연번(계약)의 여러 입금 회차를 1행으로 집계 (TLO 방식)
- false: DB 행 그대로 (BRIDGE, 세부현황 방식)

## reference_filters
- 사용자가 요청파일의 대상자 명단과 마스터 DB를 대조하라고 한 경우에만 사용하세요.
- `source_file_index`: 대상자 명단이 들어 있는 요청파일 번호(1부터 시작)
- `column`: 대상값이 있는 열 번호(1-based)
- `data_start_row`: 실제 대상값이 시작되는 행 번호
- `db_indices`: 마스터 DB에서 비교할 컬럼. 요청파일 헤더와 마스터 DB 컬럼 목록을 보고 선택하세요. 예: 교직원번호 [30,34], 연구자명/발명자명 [29,33], 사업자등록번호 [9], 업체명 [3], 기술명 [28], 과제번호 [65,66].
- `match_type`: "exact" 또는 "contains". 코드/번호는 exact 또는 contains, 이름/기관명은 contains를 권장합니다.
- 대상자 명단이 아니라 작성요령만 있으면 빈 배열 [] 로 두세요.

## db_filters
- 사용자의 자연어 요청이나 작성요령의 인정/제외 기준이 마스터 DB 자체 조건이면 여기에 구조화하세요.
- `db_index`: 마스터 DB 컬럼 번호(0-based). 반드시 마스터 DB 컬럼 목록 기준으로 고르세요.
- `operator`: eq, neq, in, not_in, contains, not_contains, gt, gte, lt, lte, between, date_between, year_eq, is_empty, not_empty 중 하나.
- `value` 또는 `values`: 비교값. between/date_between은 값 2개를 `values`로 넣으세요.
- `value_type`: text, number, date, auto 중 하나. 금액/건수는 number, 날짜는 date를 권장합니다.
- 예시:
  - "현금입금액 1억원 이상" → {{"db_index": 76, "operator": "gte", "value": 100000000, "value_type": "number"}}
  - "부처명이 교육부" → {{"db_index": 57, "operator": "contains", "value": "교육부", "value_type": "text"}}
  - "기술자문 제외" → {{"db_index": 44, "operator": "not_in", "values": ["9", "기술자문"], "value_type": "text"}}
  - "계약일 20260101~20260630" → {{"db_index": 1, "operator": "date_between", "values": ["20260101", "20260630"], "value_type": "date"}}

## unit (sum 타입에서)
- "won": 원 그대로
- "million": 백만원 (1/1,000,000)
- "thousand": 천원 (1/1,000)

## data_start_row
실제 데이터를 쓸 첫 행 번호 (헤더/제목 행 다음)
"""

    result = create_json_response(
        api_key=api_key,
        prompt=prompt,
        instructions=(
            "너는 기술이전 마스터 DB와 제출 양식 사이의 컬럼 매핑을 설계하는 전문가다. "
            "반드시 유효한 JSON 객체만 출력한다."
        ),
        max_output_tokens=4000,
    )

    # 매핑 결과 로그 출력 (디버깅용)
    print(f"  📋 AI 매핑 결과:")
    print(f"     시트: {result.get('target_sheet')} | 데이터시작행: {result.get('data_start_row')} | 필터: {result.get('filter_mode')} | 집계: {result.get('group_by_contract')}")
    for col in result.get('columns', []):
        if col.get('type') == 'db_value':
            print(f"     열{col['col']} [{col.get('label','')}] ← DB[{col['db_index']}] ({col.get('transform','')})")
        else:
            print(f"     열{col['col']} [{col.get('label','')}] ← {col['type']} {col.get('value', col.get('db_indices', ''))}")
    print(f"  📄 매핑 JSON 전체:\n{json.dumps(result, ensure_ascii=False, indent=2)}")
    return result


def get_contract_type(row):
    """계약형태 자동 판단"""
    trade     = str(row[44]).strip() if row[44] is not None else ""
    has_fixed = safe_int(row[52]) > 0
    has_cac   = bool(str(row[51]).strip()) if row[51] else False
    if trade in ("3", "8", "노하우", "단위연구노하우"):
        return "③"
    if has_fixed and has_cac:
        return "①+②"
    if has_fixed:
        return "①"
    if has_cac:
        return "②"
    return "①"


def aggregate_by_contract(rows, year):
    """같은 연번의 여러 입금 회차를 1행으로 집계."""
    contracts = {}
    for row in rows:
        seq = str(row[0]) if row[0] else None
        if not seq:
            continue
        payment = safe_int(row[76])
        p_year  = get_year(row[73])
        if seq not in contracts:
            contracts[seq] = {"row": row, "total_pay": 0, "pay_years": set()}
        contracts[seq]["total_pay"] += payment
        if p_year:
            contracts[seq]["pay_years"].add(p_year)

    result = []
    for seq, info in contracts.items():
        r     = info["row"]
        pay   = info["total_pay"]
        years = sorted(info["pay_years"])
        py    = year if year in years else (years[-1] if years else year)
        # aggregate row에 total_pay 저장 (col 76 override)
        row_list = list(r)
        # 입금 집계값을 별도 속성으로 전달할 수 없으므로 tuple에 붙이기
        result.append((r, pay, py))

    result.sort(key=lambda x: x[0][1]
                if isinstance(x[0][1], (datetime.datetime, datetime.date))
                else datetime.datetime(1900, 1, 1))
    return result


def patch_missing_columns(mapping, ws, db_columns_text, data_start):
    """
    AI가 헤더 열을 누락했을 때 db_columns.md 키워드로 자동 보완.
    헤더 행(data_start - 1)을 읽어 매핑된 열 번호와 비교한다.
    """
    # 기관마다 표현이 다를 수 있는 별칭 → db_index 직접 매핑
    ALIASES = {
        "계약명": 28, "이전기술명": 28, "기술이름": 28, "기술 명칭": 28,
        "발명자": 29, "연구책임자": 29, "발명인": 29,
        "계약업체": 3,  "도입기업": 3, "기업명": 3,
        "계약일": 1, "이전계약일": 1,
        "입금일": 73, "납입일": 73,
    }

    # db_columns.md 파싱: | db_index | label | desc |
    db_lookup = {}  # 키워드 → db_index
    for line in db_columns_text.splitlines():
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if len(parts) >= 2:
            try:
                idx = int(parts[0])
            except ValueError:
                continue
            label = parts[1]
            # 숫자 접두어 제거: "27.기술명" → "기술명"
            keyword = label.split(".")[-1].strip() if "." in label else label
            if keyword:
                db_lookup[keyword] = idx
    # 별칭 병합
    db_lookup.update(ALIASES)

    header_row = data_start - 1
    mapped_cols = {c["col"] for c in mapping.get("columns", [])}

    added = []
    for col_idx in range(1, (ws.max_column or 0) + 1):
        if col_idx in mapped_cols:
            continue
        header_val = ws.cell(header_row, col_idx).value
        if not header_val:
            continue
        # 헤더 텍스트에서 키워드 매칭
        header_clean = str(header_val).replace("\n", " ").strip()
        matched_idx = None
        for keyword, db_idx in db_lookup.items():
            if keyword and keyword in header_clean:
                matched_idx = db_idx
                break
        if matched_idx is not None:
            mapping.setdefault("columns", []).append({
                "col": col_idx,
                "type": "db_value",
                "db_index": matched_idx,
                "transform": "str",
                "label": header_clean[:30],
                "_auto_patched": True,
            })
            added.append(f"열{col_idx}[{header_clean[:20]}]→DB[{matched_idx}]")
        else:
            print(f"  ⚠️ 열{col_idx} [{header_clean[:20]}] — 매핑 없음, 수동 확인 필요")

    if added:
        print(f"  🔧 자동 보완된 열: {', '.join(added)}")
    return mapping


def get_header_cells(ws, data_start):
    """데이터 시작행 바로 위를 헤더 행으로 보고 비어 있지 않은 헤더를 수집한다."""
    header_row = max(1, int(data_start or 4) - 1)
    headers = {}
    for col_idx in range(1, (ws.max_column or 0) + 1):
        value = ws.cell(header_row, col_idx).value
        if value is not None and str(value).strip():
            headers[col_idx] = str(value).replace("\n", " ").strip()
    return header_row, headers


def validate_mapping(mapping, ws, data_start):
    """
    AI 매핑을 실행 전에 검증한다.
    누락 헤더, 낮은 confidence, 지원하지 않는 타입, 범위 밖 열을 찾아 재매핑 루프의 입력으로 쓴다.
    """
    supported_types = {
        "sequence", "db_value", "constant", "sum", "payment_amount",
        "researcher_concat", "company_region", "payment_year",
        "large_transfer", "contract_type",
    }
    columns = mapping.get("columns", []) or []
    mapped_cols = {c.get("col") for c in columns if isinstance(c.get("col"), int)}
    header_row, headers = get_header_cells(ws, data_start)
    issues = []
    low_confidence = []

    if mapping.get("form_type") == "researcher_list":
        return {
            "ok": True,
            "retry_recommended": False,
            "header_row": header_row,
            "issues": [],
            "low_confidence": [],
            "mapped_columns": 0,
            "header_columns": len(headers),
            "overall_confidence": mapping.get("confidence", {}).get("overall"),
        }

    if not columns:
        issues.append({"severity": "error", "message": "columns 배열이 비어 있습니다."})

    target_sheet = mapping.get("target_sheet")
    if target_sheet and target_sheet != ws.title:
        issues.append({
            "severity": "warning",
            "message": f"target_sheet '{target_sheet}'가 실제 선택 시트 '{ws.title}'와 다릅니다.",
        })

    for col_def in columns:
        col = col_def.get("col")
        col_type = col_def.get("type", "db_value")
        label = col_def.get("label", "")
        confidence = col_def.get("confidence")

        if not isinstance(col, int) or col < 1 or col > (ws.max_column or 0):
            issues.append({
                "severity": "error",
                "message": f"열 번호가 유효하지 않습니다: {col} ({label})",
            })
        if col_type not in supported_types:
            issues.append({
                "severity": "error",
                "message": f"지원하지 않는 column type입니다: {col_type} ({label})",
            })
        if col_type == "db_value" and not isinstance(col_def.get("db_index"), int):
            issues.append({
                "severity": "error",
                "message": f"db_value 타입인데 db_index가 없습니다: 열{col} ({label})",
            })
        if col_type == "sum" and not col_def.get("db_indices"):
            issues.append({
                "severity": "error",
                "message": f"sum 타입인데 db_indices가 없습니다: 열{col} ({label})",
            })
        if isinstance(confidence, (int, float)) and confidence < 0.70:
            low_confidence.append({
                "col": col,
                "label": label,
                "confidence": confidence,
                "evidence": col_def.get("evidence", ""),
            })

    missing_headers = []
    for col_idx, header in headers.items():
        if col_idx not in mapped_cols and not header.startswith(("비고", "작성자", "확인")):
            missing_headers.append({
                "col": col_idx,
                "letter": get_column_letter(col_idx),
                "header": header[:80],
            })

    if missing_headers:
        issues.append({
            "severity": "warning",
            "message": f"헤더가 있으나 매핑되지 않은 열 {len(missing_headers)}개가 있습니다.",
            "missing_headers": missing_headers[:30],
        })

    overall = mapping.get("confidence", {}).get("overall")
    retry_recommended = bool(
        any(i["severity"] == "error" for i in issues)
        or missing_headers
        or (isinstance(overall, (int, float)) and overall < 0.75)
    )

    return {
        "ok": not issues and not low_confidence,
        "retry_recommended": retry_recommended,
        "header_row": header_row,
        "issues": issues,
        "low_confidence": low_confidence,
        "mapped_columns": len(mapped_cols),
        "header_columns": len(headers),
        "overall_confidence": overall,
    }


def add_low_confidence_manual_items(mapping, validation):
    manual = mapping.setdefault("manual_inputs", [])
    for item in validation.get("low_confidence", []):
        text = (
            f"열{item.get('col')} {item.get('label', '')} — "
            f"AI 확신도 {item.get('confidence')}, 근거 확인 필요"
        )
        if text not in manual:
            manual.append(text)
    return mapping


def apply_mapping(ws, rows_data, mapping, data_start, group_by_contract, year):
    """매핑 JSON에 따라 데이터를 채운다."""
    columns = mapping.get("columns", [])
    total_income = 0

    for row_idx, row_entry in enumerate(rows_data, 1):
        r = data_start + row_idx - 1

        # group_by_contract 모드면 (row, pay_amt, pay_year) 튜플
        if group_by_contract:
            row, pay_amt, pay_year = row_entry
        else:
            row = row_entry
            pay_amt  = safe_int(row[76])
            pay_year = get_year(row[73]) or year

        for col_def in columns:
            col      = col_def.get("col")
            col_type = col_def.get("type", "db_value")

            try:
                value = None

                if col_type == "sequence":
                    value = row_idx

                elif col_type == "constant":
                    value = col_def.get("value", "")

                elif col_type == "db_value":
                    db_idx    = col_def.get("db_index", 0)
                    transform = col_def.get("transform", "str")
                    if transform in TRANSFORMS:
                        value = TRANSFORMS[transform](row, db_idx)
                    else:
                        value = row[db_idx] if db_idx < len(row) else ""
                    if transform == "date_obj" and isinstance(value, (datetime.datetime, datetime.date)):
                        ws.cell(r, col).number_format = "YYYY-MM-DD"

                elif col_type == "sum":
                    indices = col_def.get("db_indices", [])
                    unit    = col_def.get("unit", "won")
                    total   = sum(safe_int(row[i]) for i in indices if i < len(row))
                    if unit == "million":
                        value = round(total / 1_000_000, 1) if total else None
                    elif unit == "thousand":
                        value = total // 1_000 if total else None
                    else:
                        value = total if total else None

                elif col_type == "researcher_concat":
                    researcher = str(row[29]).strip() if row[29] else ""
                    value = f"부산대학교({researcher})" if researcher else "부산대학교산학협력단"

                elif col_type == "company_region":
                    company    = str(row[3]).strip() if row[3] else ""
                    is_foreign = str(row[6]).strip() in ("2", "국외")
                    region     = "해외" if is_foreign else get_region_name(row[8])
                    value      = f"{company}({region})" if region else company

                elif col_type == "payment_amount":
                    unit = col_def.get("unit", "won")
                    if unit == "million":
                        value = round(pay_amt / 1_000_000, 1) if pay_amt else None
                    elif unit == "thousand":
                        value = pay_amt // 1_000 if pay_amt else None
                    else:
                        value = pay_amt if pay_amt else None

                elif col_type == "payment_year":
                    value = str(pay_year) if pay_year else ""

                elif col_type == "large_transfer":
                    value = "Y" if pay_amt >= 100_000_000 else "N"

                elif col_type == "contract_type":
                    value = get_contract_type(row)

                if value == "":
                    value = None
                ws.cell(r, col, value)

            except Exception as e:
                print(f"  ⚠️ 행{row_idx} 열{col} ({col_def.get('label','')}) 처리 오류: {e}")

        total_income += pay_amt

    return total_income


def run(master_path, template_path, year, output_path, notes="", hint="", api_key=None,
        request_file_paths=None, request_file_names=None):
    print(f"📂 마스터 DB 로딩: {master_path}")
    all_rows = load_master_db(master_path)
    request_file_paths = request_file_paths or [template_path]
    request_file_names = request_file_names or [os.path.basename(p) for p in request_file_paths]

    print(f"🔍 요청파일 전체 분석 중...")
    request_analyses = []
    request_context_parts = []
    for idx, path in enumerate(request_file_paths, 1):
        analysis = analyze(path)
        analysis["_filename"] = request_file_names[idx - 1] if idx - 1 < len(request_file_names) else os.path.basename(path)
        request_analyses.append(analysis)
        sheets_summary = [f"{s['name']}({len(s.get('rows',[]))}행)" for s in analysis["sheets"]]
        print(f"  → 요청파일 {idx}: {analysis['_filename']} / 시트: {sheets_summary}")
        request_context_parts.append(format_analysis_for_prompt(analysis, f"요청파일 {idx}: {analysis['_filename']}"))
    template_analysis = request_analyses[0]
    request_files_context = "\n".join(request_context_parts)

    ref_dir = os.path.join(os.path.dirname(__file__), "..", "references")
    with open(os.path.join(ref_dir, "db_columns.md"), encoding="utf-8") as f:
        db_columns_text = f.read()
    with open(os.path.join(ref_dir, "transform_rules.md"), encoding="utf-8") as f:
        transform_rules_text = f.read()

    # 캐시 조회
    cache = load_cache()
    cache_key, cached_mapping = find_cached_mapping(os.path.basename(request_file_paths[0]), cache)
    if cache_key:
        print(f"  📦 캐시 히트: '{cache_key}' 매핑을 참고 예시로 사용")
    else:
        print(f"  📦 캐시 없음: AI가 양식을 처음부터 분석합니다")

    print(f"🤖 AI가 요청파일 역할과 양식 지침을 분석하고 매핑 중... (10~30초 소요)")
    mapping = get_mapping_from_openai(
        template_analysis, db_columns_text, transform_rules_text,
        notes, year, hint, api_key, cached_mapping=cached_mapping,
        request_files_context=request_files_context,
        request_file_count=len(request_file_paths),
    )
    if mapping.get("notes"):
        print(f"  ℹ️  {mapping['notes']}")

    # ── 연구자 명단 특수 처리 ──────────────────────────────────────
    if mapping.get("form_type") == "researcher_list":
        print(f"  → 연구자 명단 파일 감지: 자유 양식 추출로 전환")
        source_index = int(mapping.get("output_file_index") or mapping.get("source_file_index") or 1)
        source_index = max(1, min(source_index, len(request_file_paths)))
        return run_free_format(all_rows, request_file_paths[source_index - 1], mapping, year, output_path)

    print(f"  → 매핑 완료: {len(mapping.get('columns', []))}개 컬럼")

    # 워크북 로드 및 1차 매핑 검증
    output_file_index = int(mapping.get("output_file_index") or 1)
    output_file_index = max(1, min(output_file_index, len(request_file_paths)))
    output_template_path = request_file_paths[output_file_index - 1]
    print(f"  → 최종 작성양식: 요청파일 {output_file_index} ({request_file_names[output_file_index - 1]})")
    wb = load_workbook(output_template_path)
    target_sheet = mapping.get("target_sheet", wb.sheetnames[0])
    ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.active
    data_start = mapping.get("data_start_row", 4)

    mapping_validation = validate_mapping(mapping, ws, data_start)
    if mapping_validation.get("retry_recommended"):
        print("  🔁 자동 검증에서 보완 필요 항목 발견 → AI 재매핑 1회 실행")
        print(f"     검증 결과: {json.dumps(mapping_validation, ensure_ascii=False)[:1000]}")
        mapping = get_mapping_from_openai(
            template_analysis, db_columns_text, transform_rules_text,
            notes, year, hint, api_key,
            cached_mapping=cached_mapping,
            previous_mapping=mapping,
            validation_feedback=mapping_validation,
            request_files_context=request_files_context,
            request_file_count=len(request_file_paths),
        )
        if mapping.get("form_type") == "researcher_list":
            print(f"  → 재매핑 결과 연구자 명단 파일 감지: 자유 양식 추출로 전환")
            source_index = int(mapping.get("output_file_index") or mapping.get("source_file_index") or 1)
            source_index = max(1, min(source_index, len(request_file_paths)))
            return run_free_format(all_rows, request_file_paths[source_index - 1], mapping, year, output_path)
        output_file_index = int(mapping.get("output_file_index") or output_file_index)
        output_file_index = max(1, min(output_file_index, len(request_file_paths)))
        output_template_path = request_file_paths[output_file_index - 1]
        wb = load_workbook(output_template_path)
        target_sheet = mapping.get("target_sheet", wb.sheetnames[0])
        ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.active
        data_start = mapping.get("data_start_row", 4)
        mapping_validation = validate_mapping(mapping, ws, data_start)

    mapping = add_low_confidence_manual_items(mapping, mapping_validation)
    confidence = mapping.get("confidence", {}).get("overall")
    if confidence is not None:
        print(f"  🎯 매핑 전체 confidence: {confidence}")
    if mapping_validation.get("issues"):
        print(f"  ⚠️ 매핑 검증 잔여 이슈: {len(mapping_validation['issues'])}건")

    # ── notes에서 날짜 범위·연구자 파싱 ─────────────────────────────
    notes_filters = parse_notes_filters(notes)

    # 필터링 (날짜 범위 우선, 없으면 연도 전체)
    filter_mode = notes_filters["date_mode"] if notes_filters["date_start"] \
                  else mapping.get("filter_mode", "both")

    if notes_filters["date_start"] and notes_filters["date_end"]:
        from extract_columns import filter_by_date_range
        filtered = filter_by_date_range(
            all_rows,
            notes_filters["date_start"], notes_filters["date_end"],
            contract_col=1, payment_col=73, mode=filter_mode
        )
        print(f"  → 날짜범위 필터 {notes_filters['date_start']}~{notes_filters['date_end']} ({filter_mode}): {len(filtered)}건")
    else:
        filtered = filter_by_year(all_rows, year, mode=filter_mode)
        print(f"  → {year}년 해당 행: {len(filtered)}건 (기준: {filter_mode})")

    # 연구자 필터
    if notes_filters["researchers"]:
        from extract_columns import filter_by_researchers
        before = len(filtered)
        filtered = filter_by_researchers(filtered, notes_filters["researchers"],
                                         inventor_col=29, co_inventor_col=33)
        print(f"  → 연구자 필터 {notes_filters['researchers']}: {before}건 → {len(filtered)}건")

    reference_filters = mapping.get("reference_filters") or []
    if not reference_filters:
        reference_filters = infer_reference_filters(request_file_paths, notes)
    if reference_filters:
        filtered = apply_reference_filters(filtered, request_file_paths, reference_filters)

    db_filters = mapping.get("db_filters") or []
    if db_filters:
        filtered = apply_db_filters(filtered, db_filters)

    if mapping.get("exclude_consulting", True):
        filtered = [r for r in filtered if get_bridge_type(r[37], r[44]) is not None]
        print(f"  → 기술자문 제외 후: {len(filtered)}건")

    # 집계 방식 결정
    group_by = mapping.get("group_by_contract", False)
    if group_by:
        rows_data = aggregate_by_contract(filtered, year)
        print(f"  → 계약 집계 후: {len(rows_data)}건")
    else:
        # 계약일 기준 정렬
        filtered.sort(key=lambda r: r[1] if isinstance(r[1], (datetime.datetime, datetime.date))
                      else datetime.datetime(1900, 1, 1))
        rows_data = filtered

    if ws.max_row >= data_start:
        rows_to_del = ws.max_row - data_start + 1
        if rows_to_del > 0:
            ws.delete_rows(data_start, rows_to_del)

    # AI 매핑 후처리: 헤더가 있는데 매핑 누락된 열 자동 보완
    mapping = patch_missing_columns(mapping, ws, db_columns_text, data_start)
    final_mapping_validation = validate_mapping(mapping, ws, data_start)
    mapping = add_low_confidence_manual_items(mapping, final_mapping_validation)

    print(f"📝 데이터 채우기 중... (시트: {ws.title}, {data_start}행부터)")
    total_income = apply_mapping(ws, rows_data, mapping, data_start, group_by, year)

    # 수동 입력 필요 안내
    manual = mapping.get("manual_inputs", [])
    if manual:
        print(f"\n⚠️  수동 입력 필요:")
        for m in manual:
            print(f"   • {m}")

    print(f"\n📊 검토 결과 분석 중...")
    # group_by=True면 (row, pay_amt, pay_year) 3-tuple에서 row만 추출
    # group_by=False면 rows_data 자체가 row 리스트
    raw_rows = [entry[0] for entry in rows_data] if group_by else list(rows_data)
    write_feedback_sheet(
        wb, raw_rows, year, "AI 자동 추출", total_income,
        mapping_quality=final_mapping_validation,
    )

    wb.save(output_path)
    count = len(rows_data)
    print(f"\n✅ 저장 완료: {output_path}")
    print(f"   총 {count}건 | 총수입: {total_income:,}원")
    print(f"📋 검토 결과 시트에서 이상 데이터를 확인하세요.")

    return {
        "count": count,
        "total_income": total_income,
        "mapping_notes": mapping.get("notes", ""),
        "manual_inputs": manual,
        "filter_mode": filter_mode,
        "mapping": mapping,          # 캐시 저장용
        "cache_key": _normalize_key(os.path.basename(output_template_path)),
        "mapping_quality": final_mapping_validation,
        "output_file_index": output_file_index,
    }
