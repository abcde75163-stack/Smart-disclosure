"""
기술이전 추출 결과 검토 피드백 생성 모듈

각 extract_*.py 스크립트 완료 후 호출하여
출력 엑셀 파일에 '검토 결과' 시트를 추가한다.
"""
import datetime
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

# ── 스타일 정의 ─────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

STYLE = {
    'title':    (_font(bold=True, size=13, color='1F3864'), None,           _align('left')),
    'section':  (_font(bold=True, size=10, color='FFFFFF'), _fill('2F5496'), _align('left')),
    'header':   (_font(bold=True, size=9,  color='1F3864'), _fill('D9E1F2'), _align('center')),
    'good':     (_font(size=9,    color='375623'), _fill('E2EFDA'), _align('left')),
    'warn':     (_font(size=9,    color='7F4D00'), _fill('FFF2CC'), _align('left')),
    'error':    (_font(size=9,    color='7F0000'), _fill('FFE0E0'), _align('left')),
    'normal':   (_font(size=9),   None,                              _align('left', wrap=True)),
    'number':   (_font(size=9),   None,                              _align('right')),
}

def _write(ws, row, col, value, style_key='normal', width=None):
    cell = ws.cell(row=row, column=col, value=value)
    f, fill, align = STYLE[style_key]
    cell.font      = f
    cell.alignment = align
    if fill: cell.fill = fill
    cell.border = _border()
    return cell


# ── 피드백 데이터 수집 ─────────────────────────────────────────
def analyze_extraction(data_rows, year, format_name):
    """
    추출된 데이터(리스트)를 분석하여 피드백 항목 생성
    data_rows: extract_*.py 에서 처리된 row 리스트 (원본 DB row tuple)
    """
    issues        = []   # (심각도, 연번, 항목, 내용)
    manual_items  = []   # 수동 입력 필요 항목
    stats         = {}

    total = len(data_rows)
    missing_tech_name  = 0
    missing_company    = 0
    missing_payment    = 0
    zero_income        = 0
    unknown_region     = 0
    unknown_trade      = 0
    future_contract    = 0
    large_deals        = []  # 1억 이상

    from common import (safe_int, get_year, get_region_name,
                        get_bridge_type, REGION_MAP)
    import re

    for row in data_rows:
        seq   = str(row[0]) if row[0] else '?'
        cdate = row[1]
        pdate = row[73]

        # ① 기술명 누락
        if not row[28]:
            missing_tech_name += 1
            issues.append(('오류', seq, '기술명', '기술명이 비어 있습니다. 확인 필요'))

        # ② 업체명 누락
        if not row[3]:
            missing_company += 1
            issues.append(('오류', seq, '업체명', '기술도입 업체명이 없습니다'))

        # ③ 입금일 없는데 수입료 있음
        income = safe_int(row[76])
        if income > 0 and not pdate:
            issues.append(('경고', seq, '입금일', f'수입료({income:,}원)가 있으나 입금일이 비어 있습니다'))

        # ④ 수입료 0원 (계약은 있지만 입금 없는 건)
        if pdate and income == 0:
            zero_income += 1
            issues.append(('참고', seq, '수입료', '입금일은 있으나 현금입금액이 0원입니다'))

        # ⑤ 지역코드 매핑 실패
        region_raw = str(row[8]).strip() if row[8] else ''
        if region_raw and str(row[6]).strip() not in ('2','국외'):
            m = re.match(r'^(\d+)', region_raw)
            if m:
                num = m.group(1)
                if num not in REGION_MAP and num.lstrip('0') not in REGION_MAP:
                    unknown_region += 1
                    issues.append(('경고', seq, '지역코드', f"알 수 없는 지역코드: '{region_raw}'"))

        # ⑥ 거래유형 불명확
        btype = get_bridge_type(row[37], row[44])
        if btype == '기타':
            unknown_trade += 1
            issues.append(('참고', seq, '거래유형',
                           f"거래유형 '{row[44]}' / 기술유형 '{row[37]}'가 분류표에 없어 '기타'로 처리됨"))

        # ⑦ 미래 계약일
        cyear = get_year(cdate)
        if cyear and cyear > year:
            future_contract += 1
            issues.append(('경고', seq, '계약일', f'계약일이 {year}년 이후({cyear}년)입니다. 확인 필요'))

        # ⑧ 중대형 (계약 기준 정액기술료 1억+)
        fixed = safe_int(row[52])
        if fixed >= 100_000_000:
            tech_nm = str(row[28])[:30] if row[28] else '(기술명 없음)'
            company = str(row[3])[:20]  if row[3]  else '(업체 없음)'
            large_deals.append((seq, tech_nm, company, fixed))

    # 수동 입력 필요 항목 (양식 공통)
    if format_name == 'BRIDGE 3.0':
        manual_items = [
            ('AI 기술 여부 (N열)',     'DB에 AI 분류 정보 없음 — 담당자 직접 입력'),
            ('AI 기술분류 (O열)',      'AI 기술에 해당하는 건: 1~4 분류 번호 입력'),
            ('누적중대형 여부',        '경상기술료 누적액이 1억 이상인 건 별도 확인 필요'),
        ]
    elif format_name == 'TLO혁신형':
        manual_items = [
            ('주관기관명(TMC)',  '기본값 "과학기술사업화진흥원" — 실제 TMC 확인 후 수정'),
            ('총기술료 계약액', '선급+정액 합산 자동 계산 — 계약서와 대조 확인 권장'),
        ]
    elif format_name == '기술이전 실적 세부 현황':
        manual_items = [
            ('제한사항 (V열)',  '기본값 "없음" — 계약서 기준 개별 확인 필요'),
            ('기술가치평가여부', 'DB 값 그대로 사용 — 변경사항 있으면 수정'),
        ]
    elif format_name == '전체 기술거래실적':
        manual_items = [
            ('기술분류 6자리 코드', '산업기술분류표 소분류 코드 — 수동 입력 필요'),
            ('중개기관명',         '중개 거래의 경우 기관명 확인 및 입력 필요'),
            ('NTB 등록번호',       '미등록 건은 "무"로 자동 입력됨 — 변경사항 확인'),
        ]
    else:
        manual_items = [('수동 확인 필요 항목', 'AI 판단 불가 항목은 담당자 직접 확인 필요')]

    stats = {
        '총 추출 건수':          total,
        '기술명 누락':           missing_tech_name,
        '업체명 누락':           missing_company,
        '입금액 0원 건':         zero_income,
        '지역코드 매핑 불가':    unknown_region,
        '거래유형 미분류 건':    unknown_trade,
        '중대형 기술이전 건':    len(large_deals),
    }

    return stats, issues, manual_items, large_deals


# ── 피드백 시트 작성 ───────────────────────────────────────────
def write_feedback_sheet(wb, data_rows, year, format_name, total_income=0):
    """
    워크북에 '검토 결과' 시트를 추가하고 피드백을 기록한다.
    wb: 이미 열려 있는 openpyxl Workbook
    """
    # 시트 중복 방지
    if '검토 결과' in wb.sheetnames:
        del wb['검토 결과']
    ws = wb.create_sheet('검토 결과')

    stats, issues, manual_items, large_deals = \
        analyze_extraction(data_rows, year, format_name)

    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    row = 1

    # ── 제목 ────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, f'📋 검토 결과 보고서 — {format_name} ({year}년)', 'title')
    ws.row_dimensions[row].height = 24
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, f'자동 생성일시: {now_str}  |  검토 대상: {year}년 기술이전 데이터', 'normal')
    row += 2

    # ── 1. 추출 요약 ─────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, '1. 추출 요약', 'section')
    ws.row_dimensions[row].height = 18
    row += 1

    summary_data = [
        ('총 추출 건수',    stats['총 추출 건수'],   '건',  ''),
        ('총 기술료 수입',  total_income,            '원',  f'{total_income:,}'),
        ('중대형 기술이전', stats['중대형 기술이전 건'], '건', ''),
        ('기술명 누락',     stats['기술명 누락'],     '건', '⚠️ 확인 필요' if stats['기술명 누락'] else ''),
        ('업체명 누락',     stats['업체명 누락'],     '건', '⚠️ 확인 필요' if stats['업체명 누락'] else ''),
        ('입금액 0원 건',   stats['입금액 0원 건'],   '건', '참고'),
        ('거래유형 미분류', stats['거래유형 미분류 건'], '건', '⚠️ 기타 처리됨' if stats['거래유형 미분류 건'] else ''),
    ]

    _write(ws, row, 1, '항목',   'header'); _write(ws, row, 2, '건수/금액', 'header')
    _write(ws, row, 3, '단위',   'header'); _write(ws, row, 4, '비고',      'header')
    row += 1
    for label, val, unit, note in summary_data:
        style = 'warn' if ('⚠️' in note) else 'normal'
        _write(ws, row, 1, label, style)
        _write(ws, row, 2, val,   'number')
        _write(ws, row, 3, unit,  'normal')
        _write(ws, row, 4, note,  style)
        row += 1
    row += 1

    # ── 2. 중대형 기술이전 목록 ──────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, '2. 중대형 기술이전 건 목록 (정액기술료 1억원 이상)', 'section')
    ws.row_dimensions[row].height = 18
    row += 1

    if large_deals:
        for h, c in zip(['연번','기술명','업체명','정액기술료(원)'], [1,2,4,7]):
            _write(ws, row, c, h, 'header')
        row += 1
        for seq, tech, comp, amt in large_deals:
            _write(ws, row, 1, seq,  'normal')
            _write(ws, row, 2, tech, 'normal')
            _write(ws, row, 4, comp, 'normal')
            _write(ws, row, 7, amt,  'number')
            ws.cell(row, 7).number_format = '#,##0'
            row += 1
    else:
        ws.merge_cells(f'A{row}:H{row}')
        _write(ws, row, 1, '중대형 기술이전 건 없음', 'good')
        row += 1
    row += 1

    # ── 3. 수동 입력 필요 항목 ──────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, '3. 수동 입력 / 확인 필요 항목', 'section')
    ws.row_dimensions[row].height = 18
    row += 1
    _write(ws, row, 1, '항목명', 'header')
    ws.merge_cells(f'B{row}:H{row}')
    _write(ws, row, 2, '사유 및 조치 방법', 'header')
    row += 1
    for item, reason in manual_items:
        _write(ws, row, 1, item,   'warn')
        ws.merge_cells(f'B{row}:H{row}')
        _write(ws, row, 2, reason, 'warn')
        row += 1
    row += 1

    # ── 4. 데이터 이상 목록 ─────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _write(ws, row, 1, f'4. 데이터 이상 목록 (총 {len(issues)}건)', 'section')
    ws.row_dimensions[row].height = 18
    row += 1

    if issues:
        for h, c in zip(['심각도','연번','항목','내용'], [1,2,3,4]):
            _write(ws, row, c, h, 'header')
        ws.merge_cells(f'D{row}:H{row}')
        row += 1
        sev_style = {'오류': 'error', '경고': 'warn', '참고': 'normal'}
        for sev, seq, item, msg in issues:
            st = sev_style.get(sev, 'normal')
            _write(ws, row, 1, sev,  st)
            _write(ws, row, 2, seq,  st)
            _write(ws, row, 3, item, st)
            ws.merge_cells(f'D{row}:H{row}')
            _write(ws, row, 4, msg,  st)
            row += 1
    else:
        ws.merge_cells(f'A{row}:H{row}')
        _write(ws, row, 1, '✅ 이상 데이터 없음 — 모든 항목 정상', 'good')
        row += 1

    # ── 열 너비 설정 ─────────────────────────────────────────
    col_widths = [18, 12, 16, 30, 20, 10, 16, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "FF6B35"

    print(f"  📊 검토 결과 시트 추가 완료")
    print(f"     이상 항목: {len(issues)}건 | 수동 입력 필요: {len(manual_items)}건")

    return ws
